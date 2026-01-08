import sqlite3
from  openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border,Side
import os
from src.database.db_manager import DatabaseManager
from src.models.student import Student

class StudentService:
    def __init__(self):
        self.db = DatabaseManager()

#-----------------------------------------------------
    def add_student(self, student: Student):
        """add a new student"""
        try:
            student.validate()  # validate student data before insertion
    
            query = "INSERT INTO students VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, (
                    student.std_id, student.name, student.email, student.phone,
                    student.web_design, student.info_sec, student.comm_tech,
                    student.data_struct, student.wireless_net, student.comm_skill
                ))
                conn.commit()
            return True, "Student added successfully"
        except ValueError as e:
            return False, str(e)
        except sqlite3.IntegrityError:
            return False, "Error: Student ID already exists!"
        except Exception as e:
            return False, f"Database Error: {e}"
#------------------------------------------------------------------
    def get_all_students(self):
        query = "SELECT * FROM students"
        students = []
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query)
                rows = cursor.fetchall()
                for row in rows:
                    # تحويل الصفوف إلى كائنات Student
                    students.append(Student(*row))
        except Exception as e:
            print (f"Error getting students: {e}")
            return []
        
        return students
    
#------------------------------------------------------------
    def export_to_excel(self, filename="student_grades.xlsx"):
        """تصدير الإكسل مع تنسيق كامل (ألوان، خطوط، توسيط)"""
        students = self.get_all_students()
        
        if not students:
            return False, "No students , add students"
            
        try:
            # تحديد المسار
            current_file_path = os.path.abspath(__file__)
            services_dir = os.path.dirname(current_file_path)
            src_dir = os.path.dirname(services_dir)
            root_dir = os.path.dirname(src_dir)
            final_path = os.path.join(root_dir, filename)
            
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Grades"
            
            # --- (Styles) ---
            # تنسيق العناوين (غامق، خلفية زرقاء، نص أبيض)
            header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            data_font = Font(name="Arial", size=11)
            data_align = Alignment(horizontal="center", vertical="center")
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # --- 2. إضافة وتنسيق العناوين ---
            headers = ["ID", "Name", "Email", "Phone", "Web", "Security", 
                       "Comm Tech", "Data Struct", "Wireless", "Skills", "Average"]
            sheet.append(headers)
            
            # تطبيق تنسيق العناوين على الصف الأول
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # --- 3. إضافة وتنسيق البيانات ---
            for std in students:
                row = [
                    std.std_id, std.name, std.email, std.phone,
                    std.web_design, std.info_sec, std.comm_tech,
                    std.data_struct, std.wireless_net, std.comm_skill,
                    std.calculate_average()
                ]
                sheet.append(row)
                
                # تنسيق الصف الذي تمت إضافته للتو (نستخدم max_row بدلاً من count)
                current_row = sheet.max_row 
                for cell in sheet[current_row]:
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border

            # --- 4. توسيع الأعمدة تلقائياً ---
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = (max_length + 3)

            workbook.save(final_path)
            try:
                os.startfile(final_path)
            except:
                pass

            return True, f"Done exporting  formatting successfuly{len(students)} student"
            
        except PermissionError:
            return False, "The file is open, please close it and open it again."
        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, f"Export Error: {e}"