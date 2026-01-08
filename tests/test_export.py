import pytest
import os
from openpyxl import load_workbook
from src.services.student_service import StudentService
from src.models.student import Student
from src.database.db_manager import DatabaseManager

#--تجهيز التخزين المؤقت----
@pytest.fixture
def service():
    temp_db = DatabaseManager(":memory:")
    svc = StudentService()
    svc.db =temp_db
    return svc

#---تنظيف الملفات بعد الاختبار
def cleanup_file(filename):
    """delete excel file after test"""
    current_file = os.path.abspath(__file__)
    root_dir = os.path.dirname(os.path.dirname(os.path.dirname(current_file)))
    full_path = os.path.join(root_dir, filename)
    
    if os.path.exists(full_path):
        try:
            os.remove(full_path)
        except:
            pass
#---full path----
def get_file_path(filename):
    current_file = os.path.abspath(__file__) 
    tests_dir = os.path.dirname(current_file) # مجلد tests
    project_root = os.path.dirname(tests_dir) # مجلد Student-Grade-System-V2
    return os.path.join(project_root, filename)
    
        
#11- هل يرفض التصدير اذا كانت القائمة فارقه
def test_export_empty_list(service):
    success, msg =service.export_to_excel("test_empty.xlsx")
    assert success is False
    assert "empty" in msg or"فارغ" in msg or "No student" in msg
    
#12- هل يتم انشاء الملف؟----
def test_file_creation(service):
    filename= "test_create.xlsx"
    service.add_student(Student("7486786","Ali","a@gmail.com", "764657637",87,56,78,90,45,45))
    success , msg= service.export_to_excel(filename)
    assert success is True
    assert "Done exporting  formatting successfuly1 student" in msg or "exported" in msg

    cleanup_file(filename)
    
#13- هل العناوين تبع الملف مكتوبه(HEADER)؟
def test_excel_headers(service):
    filename="test_headers.xlsx"
    service.add_student(Student("7486786","Ali","a@gmail.com", "764657637",87,56,78,90,45,45))
    service.export_to_excel(filename)
    
    full_path = get_file_path(filename)
    workbook = load_workbook(full_path)
    sheet = workbook.active
    
    assert sheet["A1"].value == "ID"
    assert sheet["B1"].value == "Name"
    workbook.close()
    cleanup_file(filename)
    
#14-هل بيانات الطالب محفظه بشكل صحيح؟  
def test_excel_content(service):
    filename= "test_content.xlsx"
    
    service.add_student(Student("7486789","Ali","a@gmail.com", "764657637",87,56,78,90,45,45))
    service.export_to_excel(filename)
    
    # قراءة الملف
    full_path = get_file_path(filename)
    
    workbok= load_workbook(full_path)
    sheet = workbok.active
    
    assert sheet.cell(row=2, column=2).value == "Ali"
    workbok.close()
    cleanup_file(filename)

#15- التحقق من حساب المعدل
def test_excel_content(service):
    filename= "test_content.xlsx"
    
    service.add_student(Student("7486786","Ali","a@gmail.com", "764657637",100,100,100,100,100,100))
    service.export_to_excel(filename)
    
    # قراءة الملف
    full_path = get_file_path(filename)
    workbok= load_workbook(full_path)
    sheet = workbok.active
    
    assert sheet.cell(row=2, column=11).value == 100.0
    workbok.close()
    cleanup_file(filename)